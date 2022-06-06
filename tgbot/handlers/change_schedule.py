from typing import Iterable

from aiogram import Dispatcher, types
from openpyxl import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Protection
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..models.db import Event
from ..models.role import UserRole
from ..services.repository import Repo


def merge_cells_with_same_values_in_column(column_index: int, sheet: Worksheet) -> None:
    col = sheet[get_column_letter(column_index)]

    prev_cell = Cell(sheet)
    start_row = 1

    for cell in col:
        if cell.value == prev_cell.value:
            sheet.merge_cells(start_row=start_row, start_column=1, end_row=cell.row, end_column=1)
        else:
            prev_cell = cell
            start_row = cell.row


def align_column(column_index: int, alignment: Alignment, sheet: Worksheet) -> None:
    col = sheet[get_column_letter(column_index)]
    for cell in col:
        cell.alignment = alignment


def create_excel_template(
        file_name: str, grade_list: Iterable[str], event_list: Iterable[Event]):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'events'

    sheet.append([g for g in grade_list])
    sheet.insert_cols(idx=1, amount=4)

    weekdays = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье']

    for weekday in range(7):
        there_are_entries = False
        for e in event_list:
            if e.weekday == weekday:
                there_are_entries = True
                sheet.append([weekdays[e.weekday], e.name, e.start, e.end])

        if not there_are_entries:
            sheet.append([weekdays[weekday]])

    merge_cells_with_same_values_in_column(1, sheet)

    align_column(1, Alignment(horizontal='center', vertical='center'), sheet)

    sheet.protection.sheet = True
    no_protection = Protection(locked=False)
    for row in sheet[f'E2:{get_column_letter(sheet.max_column)}{sheet.max_row}']:
        for cell in row:
            cell.protection = no_protection

    workbook.save(file_name)


async def send_schedule_template(m: types.Message, repo: Repo):
    grade_list = await repo.get_grade_list()
    event_list = await repo.get_event_list()
    file_name = 'events.xlsx'
    create_excel_template(file_name, grade_list, event_list)

    file = types.InputFile(file_name, 'Расписание Шаблон.xlsx')
    await m.answer_document(file)


async def start_changing_schedule(m: types.Message):
    await m.answer('Отправь excel файл нового расписания.\n'
                   '/schedule_template - чтобы посмотреть шаблон')


def register_change_schedule(dp: Dispatcher):
    dp.register_message_handler(send_schedule_template, commands=['schedule_template'], state='*', role=UserRole.ADMIN)

