from typing import Iterable, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Protection
from openpyxl.utils.cell import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..models.db import Event, EventClarification


def merge_cells_with_same_values_in_column(column_index: int, sheet: Worksheet) -> None:
    col = sheet[get_column_letter(column_index)]

    prev_cell = Cell(sheet)
    start_row = 1

    for cell in col:
        if cell.value == prev_cell.value:
            sheet.merge_cells(start_row=start_row, start_column=1, end_row=cell.row, end_column=1)
        else:
            prev_cell = cell
            start_row = cell.row


def align_column(column_index: int, alignment: Alignment, sheet: Worksheet) -> None:
    col = sheet[get_column_letter(column_index)]
    for cell in col:
        cell.alignment = alignment


def get_column_letter_by_column_header(sheet: Worksheet, column_header: str) -> str:
    for cell in sheet[1]:
        if cell.value == column_header:
            return cell.column_letter
    raise LookupError(f'There isn\'t such column header: {column_header}')


def get_column_header_by_column_letter(sheet: Worksheet, column_letter: str) -> str:
    return sheet[f'{column_letter}1'].value


def create_events_clarification_excel_template(
        file_name: str, grade_list: Iterable[str], event_list: Iterable[Event],
        clarification_list: Iterable[EventClarification]):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'events_clarification'

    # Add grades
    sheet.append([g for g in grade_list])
    sheet.insert_cols(idx=1, amount=5)

    # Add events
    weekdays = ['Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота', 'Воскресенье']
    for e in event_list:
        sheet.append([weekdays[e.weekday], e.id, e.name, e.start, e.end])

    # Add events clarification
    for row in sheet[2:sheet.max_row]:
        row_index = row[1].row
        event_id = row[1].value
        for c in clarification_list:
            if c.event_id == event_id:
                column_letter = get_column_letter_by_column_header(sheet, c.grade)
                cell = f'{column_letter}{row_index}'
                sheet[cell] = c.clarification

    # Styling
    merge_cells_with_same_values_in_column(1, sheet)
    align_column(1, Alignment(horizontal='center', vertical='center'), sheet)

    # Cells protection
    sheet.protection.sheet = True
    no_protection = Protection(locked=False)
    for row in sheet[f'E2:{get_column_letter(sheet.max_column)}{sheet.max_row}']:
        for cell in row:
            cell.protection = no_protection

    workbook.save(file_name)


def create_events_excel_template(file_name: str, event_list: Iterable[Event]):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'events'

    sheet.append(['event_name', 'weekday', 'event_start', 'event_end'])

    for e in event_list:
        sheet.append([e.name, e.weekday, e.start, e.end])

    workbook.save(file_name)


def parse_events_clarification_excel(file_name: str) -> Tuple[EventClarification]:
    workbook = load_workbook(file_name)
    sheet = workbook['events_clarification']

    clarification_list = []
    for row in sheet[2:sheet.max_row]:
        event_id = row[1].value
        for cell in row[5:]:
            if cell.value is not None:
                grade = get_column_header_by_column_letter(sheet, cell.column_letter)
                clarification = EventClarification(event_id, grade, cell.value)
                clarification_list.append(clarification)

    return tuple(clarification_list)


def parse_events_excel(file_name: str) -> Tuple[Event]:
    workbook = load_workbook(file_name)
    sheet = workbook['events']

    event_list = []
    for row in sheet[2:sheet.max_row]:
        row = tuple(map(lambda cell: cell.value, row))
        (event_name, weekday, event_start, event_end) = row

        event_list.append(Event(177013, event_name, None, weekday, event_start, event_end))

    return tuple(event_list)
